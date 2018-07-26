import asyncio
import re
from asyncio import AbstractEventLoop
import csv
import json
import threading
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Tuple
from collections import OrderedDict, defaultdict

from aiohttp import ClientSession
from tkinter import Tk, Label, Button
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.hyperlink import Hyperlink


def filter_fields(entity) -> OrderedDict:
    files = []
    file_download_url = "http://zakupki.mos.ru/api/Core/FileStorage/GetDownload"
    for file in entity['files']:
        files.append(f"{file_download_url}/?fileHash={file['fileStorage']['fileHash']}")

    row = OrderedDict([
        ('customer', entity['company']['name']),
        ('endDate', entity['endDate']),
        ('companyCustomerRegionName', entity['companyCustomerRegionName']),
        ('name', entity['name']),
        ('startCost', entity['startCost']),
        ('maxDays', entity['maxDays']),
        ('fileLinks', '\n'.join(files))
    ])

    return row


def get_header() -> OrderedDict:
    header = OrderedDict([
        ('customer', "Заказчик"),
        ('endDate', "Дата окончания"),
        ('companyCustomerRegionName', "Регион"),
        ('name', "Название"),
        ('startCost', "Начальная стоиомость"),
        ('maxDays', "Дней на поставку"),
        ('fileLinks', "Ссылки на файлы"),
    ])

    return header


def adjust_size(sheet: Worksheet) -> None:
    column_widths = defaultdict(int)
    row_heights = defaultdict(int)
    for i, row in enumerate(sheet):
        for j, cell in enumerate(row):
            try:
                column_widths[j] = max(column_widths[j], len(max(cell.value.split('\n'))))
                row_heights[i] = max(row_heights[i], cell.value.count('\n'))
            except:
                pass

    for col_num, column in enumerate(sheet.columns):
        sheet.column_dimensions[get_column_letter(col_num+1)].width = column_widths[col_num]

    for row_num, row in enumerate(sheet.rows):
        sheet.row_dimensions[row_num+1].height = 25 # row_heights[row_num]


def is_link(value: str) -> bool:
    if re.findall(r'\bhttp:\S+', value):
        return True
    return False


def convert_to_hyperlink(value: str):
    result = []
    for link in re.findall(r'\bhttp:\S+', value):
        result.append(f'=HYPERLINK("{link}", "{link}")\n')

    return ' '.join(result)


def extract_links(value: str) -> Tuple[str, List[Hyperlink]]:
    """Преобразуем ссылки с HYPERLINK если они имеются"""
    hyperlink_list = []
    count = 0
    for word in value.split('\n'):
        if is_link(word.strip()):
            hyperlink = Hyperlink(ref=f"Файл {count+1}", location=word)
            hyperlink_list.append(hyperlink)
            count += 1

    return value, hyperlink_list


def convert_csv_to_excel(csv_file_name: Path, encoding='utf-8') -> None:
    wb = Workbook()
    sheet = wb.active

    CSV_SEPARATOR = ";"

    with open(csv_file_name, encoding=encoding) as f:
        reader = csv.reader(f, delimiter=CSV_SEPARATOR)
        for r, row in enumerate(reader):
            for idx, val in enumerate(row):
                cell = sheet.cell(row=r + 1, column=idx + 1)
                # TODO: сделать удобное отображение ссылок
                val, links = extract_links(val)
                cell.value = val

                for link in links:
                    cell.value = link.ref
                    cell.hyperlink = link.location
                    idx += 1
                    # TODO тут творится какая-то дичь. Разобраться.
                    cell = sheet.cell(row=cell.row, column=idx+1)

    adjust_size(sheet)
    wb.save(csv_file_name.stem + '.xlsx')


def save_entities_data(data: List[OrderedDict],
                       csv_file_name: Path,
                       encoding="windows-1251") -> None:
    header = get_header()
    with open(csv_file_name, "w", encoding=encoding, newline='') as file:
        writer = csv.writer(file, delimiter=';')
        writer.writerow([v for k, v in header.items()])

    with open(csv_file_name, "a", encoding=encoding, newline='') as file:
        dict_writer = csv.DictWriter(file, header, delimiter=';')
        dict_writer.writerows(data)


async def parse_entities(session: ClientSession, entity: Dict[str, str]) -> OrderedDict:
    get_entity_url = "http://zakupki.mos.ru/api/Cssp/OfferAuction/GetEntity"
    response = await session.get(get_entity_url, params={'id': entity['id']})
    raw_data = await response.text()
    entity_data = json.loads(raw_data)
    # with open("entity.json", "w") as file:
    #     file.write(raw_data)

    return filter_fields(entity_data)


async def get_all_entities(session: ClientSession) -> str:
    """
    Запрос почему-то возвращает абсолютно все котировки за всё время независимо от параметров,
    но вообще возможные значения фильтра такие:
    filter:
        stateIdIn:
            19000002 - активная
            19000003 - не состоялась
            19000004 - проведена
            19000005 - снята с публикации
    """
    api_request = "http://zakupki.mos.ru/api/Cssp/OfferAuction/PostQuery"
    payload = {"filter": {"stateIdIn": []},
               # "take": "10",
               "skip": 0,
               "order": [{"field": "id",
                          "desc": True}],
               "withCount": True}

    response = await session.post(api_request, data=payload)

    return await response.text()


async def run_parsing(label: Label):
    async with ClientSession() as session:
        index_url = "http://zakupki.mos.ru/#/offerauction"
        async with session.get(index_url) as response:
            await response.text()  # делаем запрос для получения кук

            raw_response = await get_all_entities(session)
            data = json.loads(raw_response)

            # with open('response.json', 'w', encoding="utf-8") as file:
            #     file.write(raw_response)
            #
            # with open('response.json', encoding='utf-8') as file:
            #     data = json.loads(file.read())

            """получаем активные котировки"""
            entites = [item for item in data['items'] if item['state']['name'] == "Активная"]
            total_entities = len(entites)

            entities_info = []
            csv_file_name = Path(datetime.now().strftime("%Y.%m.%d-%H.%M.csv"))

            for counter, entity in enumerate(entites):
                print(f"{counter+1}/{total_entities}", end='\r', flush=True, sep='')
                entities_info.append(await parse_entities(session, entity))
                label.config(text=f"{counter+1}/{total_entities}")
                save_entities_data(entities_info, csv_file_name)
                convert_csv_to_excel(csv_file_name, encoding='windows-1251')

            csv_file_name.unlink()
            print('')
            label.config(text=f"Завершено")
    exit()


def _asyncio_thread(async_loop: AbstractEventLoop, label) -> None:
    async_loop.run_until_complete(run_parsing(label))
    async_loop.close()


def run_in_thread(event_loop: AbstractEventLoop, label) -> None:
    label.config(text="Выполняется парсинг...")
    threading.Thread(target=_asyncio_thread, args=(event_loop, label)).start()


def start_application() -> None:
    root = Tk()
    root.title("Парсер сайта zakupki.mos.ru")

    label = Label(root, width=20, height=10)
    label.grid(row=0, column=0)

    event_loop = asyncio.get_event_loop()
    button = Button(master=root, text='Run', command=lambda: run_in_thread(event_loop, label))
    button.grid(row=0, column=1)

    root.mainloop()


if __name__ == "__main__":
    start_application()
